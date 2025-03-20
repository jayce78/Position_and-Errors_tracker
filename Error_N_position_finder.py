import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

# Define headers for your Excel file (adjust as needed)
HEADERS = ["Timestamp", "LAT", "LONG", "SPEED KTS", "CALC SPEED KTS", "COURSE", "DISTANCE SINCE LAST", "TOTAL DIST BETWEEN REPORTS"]

def process_excel(file_path, timestamps):
    try:
        # Load Excel file
        df = pd.read_excel(file_path, header=None, engine='openpyxl')

        # Assign column headers
        df.columns = HEADERS

        # Convert Timestamp column to datetime
        df["Timestamp"] = pd.to_datetime(df["Timestamp"], format='%d.%m.%Y %H:%M', errors='coerce')

        # Convert input timestamps to datetime
        timestamps = [datetime.strptime(ts, '%d.%m.%Y %H:%M') for ts in timestamps]

        matched_indices = []
        
        for ts in timestamps:
            closest_idx = (df["Timestamp"] - ts).abs().idxmin()
            if abs(df.at[closest_idx, "Timestamp"] - ts) <= timedelta(minutes=10):  # ±10 min tolerance
                matched_indices.append(closest_idx)

        df["Summed_Values"] = ""  # Ensure column 9 exists for summed values
        
        # Compute the summed values between matched positions
        for i in range(len(matched_indices) - 1):
            start_idx, end_idx = matched_indices[i], matched_indices[i + 1]
            df.at[end_idx, "Summed_Values"] = df.iloc[start_idx + 1:end_idx]["DISTANCE SINCE LAST"].sum()

        # Extract only the highlighted positions
        extracted_df = df.loc[matched_indices].copy()

        # Format timestamps before saving
        extracted_df["Timestamp"] = extracted_df["Timestamp"].dt.strftime('%d.%m.%Y %H:%M')

        # Save extracted data with headers
        output_file = os.path.join(os.path.dirname(file_path), "highlighted_positions_only.xlsx")
        extracted_df.to_excel(output_file, index=True, header=True)

        messagebox.showinfo("Success", f"Extracted file saved at:\n{output_file}")
        os.startfile(os.path.dirname(file_path))

    except Exception as e:
        messagebox.showerror("Error", str(e))


def apply_highlighting(output_file, matched_indices):
    wb = load_workbook(output_file)
    ws = wb.active
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for idx in matched_indices:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=idx + 1, column=col).fill = highlight_fill
    
    wb.save(output_file)

def select_file(entry):
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

def run_processing():
    file_path = entry_file_1.get()
    timestamps_text = text_timestamps.get("1.0", tk.END).strip()
    
    if not file_path:
        messagebox.showerror("Error", "Please select an Excel file.")
        return
    if not timestamps_text:
        messagebox.showerror("Error", "Please enter timestamps.")
        return
    
    timestamps = timestamps_text.split("\n")
    process_excel(file_path, timestamps)

def extract_datetime(text):
    pattern = r'\b\d{4}-\d{2}-\d{2} \d{2}:\d{2}\b'
    matches = re.findall(pattern, text)
    return [pd.to_datetime(match, format='%Y-%m-%d %H:%M') for match in matches]

def search_and_highlight():
    file_path = entry_file_2.get()
    raw_input = text_search.get("1.0", tk.END).strip()
    
    if not file_path:
        messagebox.showerror("Error", "Please select an Excel file.")
        return
    
    search_queries = extract_datetime(raw_input)
    if not search_queries:
        messagebox.showerror("Error", "No valid date-time values found.")
        return
    
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        df['DateTime_UTC'] = pd.to_datetime(df['Date_UTC'].astype(str) + ' ' + df['Time_UTC'].astype(str), format='%Y-%m-%d %H:%M')
        
        matches = df['DateTime_UTC'].isin(search_queries)
        
        if matches.any():
            wb = load_workbook(file_path)
            ws = wb.active
            highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            for index, match in enumerate(matches, start=2):
                if match:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=index, column=col).fill = highlight_fill
            
            wb.save(file_path)
            messagebox.showinfo("Success", "Matching rows highlighted in the Excel file.")
        else:
            messagebox.showinfo("No Matches", "No matching rows found.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def reset_fields():
    """Clears all input fields and text areas in the GUI."""
    entry_file_1.delete(0, tk.END)
    entry_file_2.delete(0, tk.END)
    text_timestamps.delete("1.0", tk.END)
    text_search.delete("1.0", tk.END)

# Create main window
root = tk.Tk()
root.title("Position and Error Processing GUI")
root.geometry("1000x400")

# Find and select events matching given timestamps
tk.Label(root, text="NavFleet Position Highlighter").grid(row=0, column=0, padx=20, pady=5)
entry_file_1 = tk.Entry(root, width=50)
entry_file_1.grid(row=1, column=0, padx=20)
tk.Button(root, text="Browse", command=lambda: select_file(entry_file_1)).grid(row=1, column=1)
tk.Label(root, text="Paste Timestamps (dd.mm.yyyy HH:mm per line):").grid(row=2, column=0, padx=20, pady=5)
text_timestamps = tk.Text(root, height=10, width=50)
text_timestamps.grid(row=3, column=0, padx=20)
tk.Button(root, text="Run", command=run_processing).grid(row=4, column=0, pady=10)

# Find and highlight errors in OVD
tk.Label(root, text="OVD Error Highlight").grid(row=0, column=2, padx=20, pady=5)
entry_file_2 = tk.Entry(root, width=50)
entry_file_2.grid(row=1, column=2, padx=20)
tk.Button(root, text="Browse", command=lambda: select_file(entry_file_2)).grid(row=1, column=3)
tk.Label(root, text="Paste Data with Date-Time Entries:").grid(row=2, column=2, padx=20, pady=5)
text_search = tk.Text(root, height=10, width=50)
text_search.grid(row=3, column=2, padx=20)
tk.Button(root, text="Search and Highlight", command=search_and_highlight).grid(row=4, column=2, pady=10)

# Add Reset Button to the GUI
tk.Button(root, text="Reset", command=reset_fields).grid(row=5, column=1, pady=10)

# Add a label at the bottom with your name
name_label = tk.Label(root, text="Developed by Jason Leeworthy", font=("Arial", 9, "italic"))
name_label.grid(row=6, column=0, columnspan=4, pady=10)  # Ensures it spans across columns at the bottom

# Run the application
root.mainloop()

